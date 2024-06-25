'''
Author: Bipal Shakya

New __init__ ++ main


'''
print ("NMEA Simulation being initiated. Please wait a moment...")


import socket, os, random, sys, pyexcel, time,ast
# from time import sleep
from threading import Thread
# from SocketServer import ThreadingMixIn
from openpyxl import load_workbook

from main import NmeaSimulator
nmeaSim = NmeaSimulator()

from readTextFile import readTextFile



# Separate routefiles listed out for the current version; until all routes can
# be combined into a single structured markup file/text-based database
routeFile                    = os.listdir('./routes_latlong')
# outefile chosen at random at the start of the script; during the scripts
# session, it will loop to and back the chosen route info
chosenRouteFile              = "./routes_latlong/" + random.choice(routeFile)
staticSentencePath           = "./staticSentences/"
file_name                    = 'NMEA_ConfigTemplate.xlsx'
absnt_ver_stat_keywords     = ["not in", "not found", "unknown", "deprecated"]

wbName                       = 'NMEA_ConfigTemplate.xlsx'
sheetName                    = 'NMEA_Config_Data'        # Sheet1=prod, Sheet2=devtest

def open_sheet():
    return_value = False
    try:
        book = pyexcel.get_book(file_name=file_name)
        global sheet
        sheet  = book.sheet_by_index(0)
    except Exception as error_message:
        print("\n** ERROR : {} \n\n".format(error_message))
        return return_value
    return_value = True
    return return_value


# This gets the sentence format according to the nmea Version.
# The given sentence version is checked with its corresponding column in
# excel sheet. The sentence version vs. excel column name is defined in
# a dictionary "switcher"

# According to the source excel sheet, column B holds all information about
# sentence ID. Therefore looping through each cell in Column B
# and checking if the user-given sentence ID matches any in Column B.
# If it does, return the intersection between the matching sentence ID's row
# and selected version's column.

def getSentenceFormat(nmeaVersion, sentenceId):
    # he information below needs to match that with Excel sheet. Need to make this
    # more dynamic later.
    def getVersionRow(nmeaVersion):
        switcher = {
            '1.5' :  'C',
            'IEC' :  'D',
            '2'   :  'E',
            '2.0' :  'E',
            '2.01':  'F',
            '2.1' :  'G',
            '2.2' :  'H',
            '2.3' :  'I',
            '3'   :  'J',
            '3.0' :  'J',
            '3.01':  'K',
            '4'   :  'L',
            '4.0' :  'L',
            '4.1' :  'M'
        }
        return switcher.get(nmeaVersion, '0')

    # 'i' is set as 2 because 1st row represents heading names; and not actual values
    lastRowNum = ws.max_row
    i = 2
    while (i <= lastRowNum):
        if sentenceId == (ws[('B'+str(i))].value):
            versionExcelRow = str(getVersionRow(nmeaVersion))

            try:
                sentenceRawFormat = ws[(versionExcelRow+str(i))].value
            except AttributeError:
                print ("Fatal Error: Invalid Version Input!")
                sys.exit(0)

            for keyword in absnt_ver_stat_keywords:
                if keyword in sentenceRawFormat.lower():
                    return '0'
                    # print "The selected variables doesn't exist for NMEA version " + nmeaVersion
                    # sys.exit(2)

            else:
                return sentenceRawFormat.replace(' ', '')

        i += 1

#staticSentesnces = readTextFile(2, staticSentenceFile)

# Multi threaded Python server : TCP Server Socket Thread Pool
class ClientThread(Thread):

    def __init__(self, ip, port, talker_id, sentense_list, nmea_version, time_to_sleep, equipment_name,replay_from_file):
        Thread.__init__(self)
        self.ip = ip
        self.port = port
        self.talker_id = talker_id
        self.sentense_list = sentense_list
        self.nmea_version = nmea_version
        self.time_to_sleep = time_to_sleep
        self.equipment_name = equipment_name
        self.replay_from_file = replay_from_file
        print ("[+] " + ip + ":" + str(port))
        if self.replay_from_file == "Y":
            self.staticSentences = readTextFile(3, staticSentencePath+str(port))
            
    def run(self):
        
        tcpServer = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        tcpServer.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        tcpServer.bind((self.ip, self.port))
        tcpServer.listen(4)
        print("accepting a connection")
        (conn,(self.ip,self.port)) = tcpServer.accept()
        self.tcpServer = tcpServer
        self.conn = conn
        def staticSentence():
            while True:
                try:
                    # TCP connection
                    
                    for line in self.staticSentences:
                        self.conn.send((line+'\n').encode("utf-8"))
                        print("print data has sent")
                        time.sleep(self.time_to_sleep)
                except Exception as errormessage:
                    print("While sending data :", errormessage)
                    (self.conn, (self.ip, self.port)) = tcpServer.accept()
                    continue
                time.sleep(0.02)
        if self.replay_from_file == "Y":
            staticSentence()


        loopCount                    = 0
        latLongInverseRead           = False
        # toFroLoopway is primarily for latitude and longitude coordinates.
        # Get number of lines in chosenRouteFile so that looping through and "line to
        # get" can be handled accordingly
        # print "\nChosen GPS Route File: " + chosenRouteFile + "\n"        # Checkpoint
        toFroLoopway = int(readTextFile(1, chosenRouteFile))

        while True:
            # print loopCount, toFroLoopway            # Checkpoint
            # The following logic to read back and forth through the route file

            if ((loopCount < (toFroLoopway)) and latLongInverseRead == False):
                loopCount += 1
            else:
                latLongInverseRead = True
                loopCount -= 1
                if loopCount == 1:
                    latLongInverseRead = False

            for sentenceId in self.sentense_list:
                if self.talker_id == 'LC':
                    sentenceId = 'LC' + sentenceId

                # sentence_log = open('.sentence_logs/' + self.equipment_name, 'a')
                # Get the raw sentence format based on version and sentenceId selected
                sentenceFormat = getSentenceFormat(self.nmea_version, sentenceId)

                MESSAGE = str(nmeaSim.main(self.talker_id, sentenceId, loopCount, chosenRouteFile, sentenceFormat, self.nmea_version)) + "\n"
                
                if ("None" in MESSAGE):
                    continue
                else:
                    try:
                        # sentence_log.write(MESSAGE)
                        conn.send(MESSAGE)
                        #print(MESSAGE,"meshhh")
                    except:
                        # sentence_log.close()
                        # os.system('cp .sentence_logs/' + self.equipment_name + ' sentence_logs/' + self.equipment_name)
                        # sentence_log = open('.sentence_logs/' + self.equipment_name, 'w')
                        # sentence_log.close()
                        (conn, (self.ip,self.port)) = tcpServer.accept()


            time.sleep(self.time_to_sleep)
    

wb = load_workbook(wbName)
try:
    ws = wb[sheetName]
except KeyError:
    print (sheetName + " does not exist!")
    sys.exit(0)

open_sheet()
row_id = 0              # start at row 0
cell_set = set()
threads = []
sheet = pyexcel.get_sheet(file_name=wbName)
for current_row in sheet.rows():
    if "NMEA" in str(sheet[row_id, 3]):
        sim_ip_addr = str(sheet[row_id, 10]).strip()
        port_num = str(sheet[row_id, 11]).strip()
        talker_id = (str(sheet[row_id, 12]).replace(" ", "")).partition(":")[0]
        sentense_list = (str(sheet[row_id, 13]).replace(" ", "")).split(",")
        nmea_version = str(sheet[row_id, 14])
        time_to_sleep = str(sheet[row_id, 15])
        equipment_name = str(sheet[row_id, 1]).replace(' ', '')
        replay_from_file = str(sheet[row_id, 16])
        if time_to_sleep == '':
            time_to_sleep = '1'

        cth = ClientThread(sim_ip_addr, int(port_num), talker_id, sentense_list, nmea_version, float(time_to_sleep), equipment_name,replay_from_file)
        
        print ("This port belongs to: {}, NMEA version {}".format(equipment_name, nmea_version))
        print ("Throughput: {} sentence(s) per {} second(s)\n".format(str(len(sentense_list)), str(time_to_sleep)))

        cth.daemon = True
        threads.append(cth)
        cth.start()

    row_id = row_id + 1
    

while True:
    time.sleep(0.02)
    continue


